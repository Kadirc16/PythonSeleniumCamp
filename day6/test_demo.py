from selenium import webdriver
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.common.by import By
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.support import expected_conditions
from selenium.webdriver.common.action_chains import ActionChains
import pytest
from pathlib import Path #testi oluşturup yaptığımız güne klasör oluştur
from datetime import date
from constants import globalConstants
import openpyxl

#prefix => ön ek test_ isimlendirmelerimizde test_ kullandığımız tüm alanları burada gösteriyor. classlar,fonksiyonlarımız dahil test_ fonksiyonu ile yazılmalı , classın içinde yardımcı fonks olduğunda bunları görmez.bu yüzden test fonksiyonlarına test_ ile başlanmalı.
#
#postfix


class Test_DemoClass:
    #her testten önce çağırılır.
    def setup_method(self):
        self.driver = webdriver.Chrome(ChromeDriverManager().install())
        self.driver.maximize_window()
        self.driver.get(globalConstants.URL)
        self.folderPath =str (date.today())
        Path(self.folderPath).mkdir(exist_ok=True)
        #günün tarihini al bu tarih ile bir klasör var mı kontrol et yoksa oluştur. testlerimizi oluşturduğumuz her an o günün tarihini alıp bu tarihle klasör oluşturup bugünün testleri için oluşturulacak ss ler bu klasörde tutulacak... mk dir make direction ... exist_ok ilgili klasör halihazırda oluşturulmuşsa tekrar bunu oluşturma
        

    #her testten sonra çağrılır.
    def teardown_method(self):
        self.driver.quit() #ilgili tarayıcıyı kapatır.
    
    # setup -> demoFunc -> tear down şeklinde çalışacak
    def test_demoFunc(self):
        text = "Hello"

        assert text == "Hello"
    @pytest.mark.skip() # annotations ile bu fonksiyon teste girmemiş oldu.
    def test_demo2(self):
        assert True


    def getData (): 
        # bunu decoratore göndereceğimiz için self parametresi vermiyoruz. test fonksiyonları içine yardımcı bir fonksiyon olarak kurduk. 
        #veriyi ele al . (örn. database, excel )  

        excelFile = openpyxl.load_workbook("data/invalid_login.xlsx") #bir excel dosyasını eklememizi sağlayan fonksiyon
        selectedSheet = excelFile["Sheet1"]
        
        totalRows = selectedSheet.max_row
        data=[]
        for i in range(2,totalRows+1): # 1.satırda username ve password vardı. 2 den başlattık.
            
            username = selectedSheet.cell(i,1).value #bir hücreyi bulmak için o satır ve sütünun excellde nereye denk geldiğini vermemiz lazım. içindeki değeri okumamız için value eklememiz lazım.for dengesinde son sayı <= değil , < olarak görür +1 ekleyerek bunu çözebiliriz.
            password = selectedSheet.cell(i,2).value
            tupleData = (username,password)
            data.append(tupleData) # her gezdiğimiz değeri exceldeki satırlara ekleyecek bir for döngüsü
        
        
        return data


    @pytest.mark.parametrize("username,password",getData()) #birden fazla deneme yapacağımız zaman parametre olarak göndermiş oldu. username password alanına iki kez liste halinde gönderdik.
    def test_invalid_login(self,username,password):
       self.waitForElementVisible((By.ID,"user-name"))
       usernameInput =self.driver.find_element(By.ID,"user-name")
       self.waitForElementVisible((By.ID,"password"),10)
       passwordInput= self.driver.find_element(By.ID,"password")
       usernameInput.send_keys(username)
       passwordInput.send_keys(password)
       loginBtn = self.driver.find_element(By.ID,"login-button")
       loginBtn.click()
       errorMessage = self.driver.find_element(By.XPATH, "//*[@id='login_button_container']/div/form/div[3]/h3")

       #self.driver.save_screenshot(self.folderPath+"/test-invalid-login.png") bu şekilde sürekli aynı isimde kaydedip üzerine yazacak
       self.driver.save_screenshot(f"{self.folderPath}/test-invalid-{username}-{password}.png")  # bu şekilde ise girdiğimiz isimşifre halinde ss alacak
       
       #testin bitiminin bir üstünde olması daha iyi olur ilgili tarayıcının testini ss alacak.
       assert  errorMessage.text == "Epic sadface: Username and password do not match any user in this service"

    def waitForElementVisible(self,locator,timeout=5): #fonksiyonda belirtmedikçe 5 sn yi default olarak ayarladık.fonksiyonda belirtilirse onu kullanır.
       WebDriverWait(self.driver,timeout).until(expected_conditions.visibility_of_element_located(locator))



""" 
        #3 A Act Arrange Assert 
# assert her şeyimizi tamamladıktan sonra test classımıza diyoruz ki durum buysa bize doğru veya yanlış old göster.
# ilgili testin bağlı old conditionu belirtir.
# Py test te __init__ kullanamıyoruz.
# Onun yerine setup_method,teardown_method kullanılır.
# py testte print bölümü yok onun yerine ilgili yere breakpoint ekleyip debug #çalıştırılabilir

magic string -- kodlamada tırnak içerisinde yazılan string verilerdir. kodlamada bunları istemeyiz. hataya çok açık en ufak bir hatayı ide nin bize bildirme imkanı yok ve bir değişiklik yapılacağı zaman tek tek hepsinde değişiklik yapmak gerekiyor. onun yerine bu verileri tutabileceğimiz değişkenler oluşturarak ;constants(sabitler) lar oluşturararak import etmek daha temiz bir kod yazmaya ve ilerde yapılacak değişikliği tek seferde hepsinde geçerli kılmayı sağlayacaktır.

dersin sonrası selenium ide de anlatılıyor. 
 send_keys = type olarak selenium ide de görüyoruz.
 selenium ide ilgili alanın görünmesini bekliyor. webDriverWait kullanmamıza gerek yok. ama tabi ki bunları export edince hata almamak için defensive sistemi selenium ide üzerinden yapacağız.
 hata mesajının üzerine sağ tık selenium ide>assert text işlemini buradan gerçekleştirebiliriz kamp 6.gün 1.14. dk
 test_invalidlogin.py dosyası
 
 """
